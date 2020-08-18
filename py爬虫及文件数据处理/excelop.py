#coding=utf-8
from openpyxl import Workbook
import os
import re


str1 = "项目名称：";        str1_1 = "项目名称:";       #str1_2 = "项目名称及"   
str2 = "供应商名称：";      str2_1 = "中标人名称：";
str2_2 = "供应商名称:";    str2_3 = "中标人名称:";
str3 = "供应商地址：";      str3_1 = "供应商地址:";
str3_2 = "中标人地址：";    str3_3 = "中标人地址:";     str3_4 = "中标单位：";
str4 = "中标金额：";        str4_1 = "中标金额:";
str5 = "联系方式：";
str6 = "联系人：";
str7 = "www";
str8 = "----------------------------------------"

fp = open("final.txt")
testbook = Workbook()
testsheet = testbook.active     #获取当前活跃的sheet,默认是第一个sheet
line = fp.readline()
maxcount = 1; pronamecount = 2; cocount = 2; addcount = 2; moneycount = 2;
contcount = 2; numcount = 2; linkcount = 2;
number = 1
testsheet.cell(1,1).value = '编号'
testsheet.cell(1,2).value = '项目名称'
testsheet.cell(1,3).value = '供应商名称'
testsheet.cell(1,4).value = '供应商地址'
testsheet.cell(1,5).value = '中标金额'
testsheet.cell(1,6).value = '联系人'
testsheet.cell(1,7).value = '联系方式'
testsheet.cell(1,8).value = '链接'

while line:
    if (line.find(str8) != -1):
        pronamecount = cocount = addcount = moneycount = contcount = numcount = linkcount = maxcount + 1

    if (line.find(str1) != -1):
        a = line.find(str1)
        print(line[a+5:],end = '')
        testsheet.cell(pronamecount,2).value = line[a+5:]
        testsheet.cell(pronamecount,1).value = number
        number += 1
        pronamecount += 1
        maxcount = pronamecount;

    if (line.find(str1_1) != -1):
        a = line.find(str1_1)
        print(line[a+5:],end = '')
        testsheet.cell(pronamecount,2).value = line[a+5:]
        testsheet.cell(pronamecount,1).value = number
        number += 1
        pronamecount += 1
        maxcount = pronamecount;

    if (line.find(str2) != -1):
        a = line.find(str2)
        print(line[a+6:],end = '')
        testsheet.cell(cocount,3).value = line[a+6:]
        cocount += 1
        if maxcount < cocount :
            maxcount = cocount

    if (line.find(str2_1)!=-1):
        a = line.find(str2_1)
        print(line[a+6:],end = '')
        testsheet.cell(cocount,3).value = line[a+6:]
        cocount += 1
        if maxcount < cocount :
            maxcount = cocount

    if (line.find(str2_2) != -1):
        a = line.find(str2_2)
        print(line[a+6:],end = '')
        testsheet.cell(cocount,3).value = line[a+6:]
        cocount += 1
        if maxcount < cocount :
            maxcount = cocount

    if (line.find(str2_3)!=-1):
        a = line.find(str2_3)
        print(line[a+6:],end = '')
        testsheet.cell(cocount,3).value = line[a+6:]
        cocount += 1
        if maxcount < cocount :
            maxcount = cocount

    if (line.find(str3) != -1):
        a = line.find(str3)
        print(line[a+6:],end = '')
        testsheet.cell(addcount,4).value = line[a+6:]
        addcount += 1
        if maxcount < addcount :
            maxcount = addcount

    if (line.find(str3_1) != -1):
        a = line.find(str3_1)
        print(line[a+6:],end = '')
        testsheet.cell(addcount,4).value = line[a+6:]
        addcount += 1
        if maxcount < addcount :
            maxcount = addcount

    if (line.find(str3_2) != -1):
        a = line.find(str3_2)
        print(line[a+6:],end = '')
        testsheet.cell(addcount,4).value = line[a+6:]
        addcount += 1
        if maxcount < addcount :
            maxcount = addcount

    if (line.find(str3_3) != -1):
        a = line.find(str3_3)
        print(line[a+6:],end = '')
        testsheet.cell(addcount,4).value = line[a+6:]
        addcount += 1
        if maxcount < addcount :
            maxcount = addcount

    if (line.find(str3_4) != -1):
        a = line.find(str3_4)
        testsheet.cell(addcount,4).value = line[a+5:]
        print(line[a+5:],end = '')
        addcount += 1
        if maxcount < addcount :
            maxcount = addcount

    if (line.find(str4) != -1):
        a = line.find(str4)
        print(line[a+5:],end = '')
        testsheet.cell(moneycount,5).value = line[a+5:]
        moneycount += 1
        if maxcount < moneycount :
            maxcount = moneycount

    if (line.find(str4_1) != -1):
        a = line.find(str4_1)
        print(line[a+5:],end = '')
        testsheet.cell(moneycount,5).value = line[a+5:]
        moneycount += 1
        if maxcount < moneycount :
            maxcount = moneycount

    if (line.find(str5) != -1):
        a = line.find(str5)
        print(line[a+5:],end = '')
        testsheet.cell(numcount,6).value = line[a+5:]
        numcount += 1
        if maxcount < numcount :
            maxcount = numcount
    
    if (line.find(str6) != -1):
        a = line.find(str6)
        print(line[a+4:],end = '')
        testsheet.cell(contcount,7).value = line[a+4:]
        contcount += 1
        if maxcount < contcount :
            maxcount = contcount

    if (line.find(str7) != -1):
       print(line,end = '')
       testsheet.cell(linkcount,8).value = line
       linkcount += 1
       if maxcount < linkcount:
           maxcount = linkcount

    line = fp.readline()
testbook.save('Excel_test.xlsx')
fp.close()